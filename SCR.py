import openpyxl

def process_excel_task(task_name, file_name):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_name)

    # Select the first sheet in the workbook
    sheet = workbook.active

    if task_name == "sum_column":
        # Sum the values in column A and return the result
        sum_value = 0
        for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True):
            sum_value += row[0]
        return f"The sum of column A is: {sum_value}"

    elif task_name == "count_rows":
        # Count the number of rows in the sheet and return the result
        row_count = sheet.max_row
        return f"The number of rows in the sheet is: {row_count}"

    else:
        return "Invalid task name. Please provide a valid task name."

# Example usage of the function
task_name = "sum_column"
file_name = "example.xlsx"
result = process_excel_task(task_name, file_name)
print(result)