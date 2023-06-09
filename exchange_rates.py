import requests
from openpyxl import load_workbook
from xml.etree import ElementTree as ET
def pull_rates():
  url = "https://services.nbg.gov.ge/Rates/Service.asmx"

  headers = {
      'Content-Type': 'text/xml; charset=utf-8',
      'SOAPAction': 'http://www.nbg.ge/GetCurrentRates',
  }

  body = """<?xml version="1.0" encoding="utf-8"?>
  <soap:Envelope xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:xsd="http://www.w3.org/2001/XMLSchema" xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/">
    <soap:Body>
      <GetCurrentRates xmlns="http://www.nbg.ge/">
        <Currencies></Currencies>
      </GetCurrentRates>
    </soap:Body>
  </soap:Envelope>"""

  response = requests.post(url, headers=headers, data=body)
  root = ET.fromstring(response.content)

  namespace = {'soap': 'http://schemas.xmlsoap.org/soap/envelope/', 'nbg': 'http://www.nbg.ge/'}

  rates = {}
  for rate in root.findall('.//nbg:GetCurrentRatesResult/nbg:CurrencyRate', namespace):
      code = rate.find('nbg:Code', namespace).text
      quantity = rate.find('nbg:Quantity', namespace).text
      rate_value = rate.find('nbg:Rate', namespace).text
      rates[code] = (quantity, rate_value)

  print(rates)

  # Load the existing workbook
  wb = load_workbook(r'C:\Users\Kchkholaria\Desktop\Project_SCR\working_files\SNG.xlsx')
  ws = wb.active

  # Write rates to specified cells
  ws['B16'] = float(rates['USD'][1]) / int(rates['USD'][0])  # USD rate value
  ws['B17'] = float(rates['EUR'][1]) / int(rates['EUR'][0])  # EUR rate value
  ws['B18'] = float(rates['RUB'][1]) / int(rates['RUB'][0])  # RUB rate value

  # Save the workbook to a file
  wb.save(r'C:\Users\Kchkholaria\Desktop\Project_SCR\working_files\SNG.xlsx')